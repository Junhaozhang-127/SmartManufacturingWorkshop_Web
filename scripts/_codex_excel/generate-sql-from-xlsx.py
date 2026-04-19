import datetime
import os
import re
import sys

import openpyxl


def sql_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("'", "''")


def sql_literal(value, defaultable: bool) -> str:
    if value is None or value == "":
        return "DEFAULT" if defaultable else "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    if isinstance(value, datetime.datetime):
        return f"'{value:%Y-%m-%d %H:%M:%S}'"
    if isinstance(value, datetime.date):
        return f"'{value:%Y-%m-%d}'"

    text = str(value).strip()
    if text == "":
        return "DEFAULT" if defaultable else "NULL"
    return f"'{sql_escape(text)}'"


def parse_table_name(a1: str) -> str | None:
    match = re.search(r"表：([^）)]+)", a1)
    return match.group(1) if match else None


def is_model_sheet(ws) -> bool:
    a1 = ws.cell(1, 1).value
    return bool(a1) and "表：" in str(a1)


def sheet_columns(ws):
    cols: list[str] = []
    metas: list[str | None] = []
    col_index = 1
    while True:
        value = ws.cell(4, col_index).value
        if value in (None, ""):
            break
        cols.append(str(value).strip())
        metas.append(ws.cell(5, col_index).value)
        col_index += 1

    defaultable: list[bool] = []
    for meta in metas:
        text = "" if meta is None else str(meta)
        defaultable.append(
            ("默认" in text)
            or ("autoincrement" in text)
            or ("自动更新时间" in text)
            or ("自动更新" in text)
        )

    return cols, defaultable


def iter_data_rows(ws, cols_len: int):
    for row_index in range(6, ws.max_row + 1):
        row = [ws.cell(row_index, col).value for col in range(1, cols_len + 1)]
        if any(value not in (None, "") for value in row):
            yield row


SEED_ORDER = [
    "SysUser",
    "SysRole",
    "OrgUnit",
    "SysUserRole",
    "MemberProfile",
    "SysDictType",
    "SysDictItem",
    "SysConfigItem",
    "WfApprovalTemplate",
    "WfApprovalTemplateNode",
    "CompCompetition",
    "CompTeam",
    "AchvAchievement",
    "AchvContributor",
    "AchvPaper",
    "IpAsset",
    "EvalScheme",
    "GovRewardPenalty",
    "AssetDevice",
    "AssetDeviceRepair",
    "FundAccount",
    "FundApplication",
    "InvConsumable",
    "InvInventoryTxn",
    "InvConsumableRequest",
    "MemberGrowthRecord",
    "MemberStageEvaluation",
    "MemberRegularization",
    "MemberOperationLog",
    "PromApplication",
    "PromAppointment",
    "SysNotification",
]


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: generate-sql-from-xlsx.py <input.xlsx> [output.sql]")
        return 2

    xlsx_path = sys.argv[1]
    output_path = (
        sys.argv[2]
        if len(sys.argv) >= 3
        else os.path.join(os.getcwd(), "docs", "db-reimport-data.sql")
    )

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    sheets: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not is_model_sheet(ws):
            continue

        cols, defaultable = sheet_columns(ws)
        rows = list(iter_data_rows(ws, len(cols)))
        if not rows:
            continue

        table_name = parse_table_name(str(ws.cell(1, 1).value))
        if not table_name:
            continue

        sheets[sheet_name] = {
            "sheet": sheet_name,
            "table": table_name,
            "cols": cols,
            "defaultable": defaultable,
            "rows": rows,
        }

    ordered = []
    for name in SEED_ORDER:
        if name in sheets:
            ordered.append(sheets.pop(name))
    for name in wb.sheetnames:
        if name in sheets:
            ordered.append(sheets[name])

    lines: list[str] = []
    lines.append(f"-- Generated from: {os.path.basename(xlsx_path)}")
    lines.append("SET NAMES utf8mb4;")
    lines.append("SET FOREIGN_KEY_CHECKS=0;")
    lines.append("START TRANSACTION;")

    for entry in ordered:
        table = entry["table"]
        cols = entry["cols"]
        defaultable = entry["defaultable"]
        rows = entry["rows"]

        lines.append("")
        lines.append(f"-- Sheet: {entry['sheet']} -> `{table}` (rows: {len(rows)})")

        col_sql = ", ".join(f"`{c}`" for c in cols)
        lines.append(f"INSERT INTO `{table}` ({col_sql}) VALUES")

        value_lines: list[str] = []
        for row in rows:
            values = [
                sql_literal(value, defaultable[index])
                for index, value in enumerate(row)
            ]
            value_lines.append(f"  ({', '.join(values)})")

        lines.append(",\n".join(value_lines) + ";")

    lines.append("COMMIT;")
    lines.append("SET FOREIGN_KEY_CHECKS=1;")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")

    total_rows = sum(len(entry["rows"]) for entry in ordered)
    print(f"Wrote: {output_path} (tables: {len(ordered)}, rows: {total_rows})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

